from datetime import datetime
from datetime import datetime
import openpyxl

class Accesso:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(filename)
        self.worksheet = self.workbook.active

    def record_access(self, societa, nominativo):
        # Trovare l'ultima riga del foglio di lavoro
        last_row = self.worksheet.max_row + 1

        # Ottenere la data e l'ora attuali
        data = datetime.today().strftime("%Y-%m-%d")
        ora = datetime.today().strftime("%H:%M:%S")

        # Controllare se l'utente è già presente nel foglio di lavoro per oggi
        for row in self.worksheet.iter_rows(min_row=2, min_col=1, max_col=5):
            if row[0].value == data and row[2].value == nominativo:
                # Se l'utente è già presente, registrare l'uscita
                for cell in self.worksheet[row[0].row]:
                    if cell.column == 3:
                        cell.value = nominativo
                        self.worksheet.cell(row=cell.row, column=5, value=ora)
                        self.workbook.save(self.filename)
                        return "Uscita registrata"
                break
        else:
            # Se l'utente non è presente, registrare l'ingresso
            self.worksheet.cell(row=last_row, column=1).value = data
            self.worksheet.cell(row=last_row, column=4).value = ora
            self.worksheet.cell(row=last_row, column=2).value = societa
            self.worksheet.cell(row=last_row, column=3).value = nominativo
            self.workbook.save(self.filename)
            return "Ingresso registrato"



# Creare un'istanza della classe Accesso
accessi = Accesso("accessi.xlsx")

azienda = input("Inserisci l'azienda: ")
nominativo = input("Inserisci il nominativo: ")

risultato = accessi.record_access(azienda, nominativo)
print(risultato)
