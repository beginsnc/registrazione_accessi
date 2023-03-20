from datetime import datetime
import openpyxl


class Accesso:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(filename)
        self.worksheet = self.workbook.active

    def record_access(self, societa, nominativo):
        last_row = self.worksheet.max_row + 1
        current_date = datetime.today().strftime("%Y-%m-%d")
        current_time = datetime.today().strftime("%H:%M:%S")

        for row in self.worksheet.iter_rows(min_row=2, min_col=1, max_col=5):
            if row[0].value == current_date and row[2].value == nominativo:
                row[4].value = current_time
                self.workbook.save(self.filename)
                return "Uscita registrata"
        else:
            self.worksheet.cell(row=last_row, column=1, value=current_date)
            self.worksheet.cell(row=last_row, column=2, value=societa)
            self.worksheet.cell(row=last_row, column=3, value=nominativo)
            self.worksheet.cell(row=last_row, column=4, value=current_time)
            self.workbook.save(self.filename)
            return "Ingresso registrato"

    def get_unregistered_exits(self):
        unregistered_exits = []
        current_date = datetime.today().strftime("%Y-%m-%d")

        for row in self.worksheet.iter_rows():
            if (row[4].value is None or row[4].value == "") and row[0].value == current_date:
                unregistered_exits.append(row[2])
        return unregistered_exits
