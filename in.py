import openpyxl
from datetime import datetime

# aprire il file excel
workbook = openpyxl.load_workbook('accessi.xlsx')

# selezionare il foglio di lavoro
worksheet = workbook.active

# ottenere i dati di input
data = datetime.today().strftime('%Y-%m-%d')
ora = datetime.today().strftime('%H:%M:%S')
nome_societa = input("Inserisci il nome della società: ")
nominativo = input("Inserisci il nominativo: ")

# aggiungere i dati all'ultima riga del foglio di lavoro
last_row = worksheet.max_row + 1
worksheet.cell(row=last_row, column=1).value = data
worksheet.cell(row=last_row, column=4).value = ora
worksheet.cell(row=last_row, column=2).value = nome_societa
worksheet.cell(row=last_row, column=3).value = nominativo

# salvare le modifiche e chiudere il file excel
workbook.save('accessi.xlsx')
