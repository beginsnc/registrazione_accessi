from datetime import datetime

import openpyxl

# aprire il file excel
workbook = openpyxl.load_workbook('accessi.xlsx')

# selezionare il foglio di lavoro
worksheet = workbook.active

# Richiesta del nominativo tramite input
nominativo = input('Inserisci il nominativo: ')

# Verifica della data di oggi
oggi = datetime.now().date().strftime('%Y-%m-%d')

# Creazione orario di uscita
ora = datetime.today().strftime('%H:%M:%S')

# Ricerca della riga con la data di oggi e il nominativo fornito
last_row = worksheet.max_row + 1

for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=5):
    if row[0].value == oggi and row[2].value == nominativo and row[4].value == False:
        # Inserisci il nominativo nella quinta casella della stessa riga
        for cell in worksheet[row[0].row]:
            if cell.column == 3:
                cell.value = nominativo
                worksheet.cell(row=cell.row, column=5, value=ora)


# Salvataggio delle modifiche al file Excel
workbook.save('accessi.xlsx')